"""
report_generator.py

Professional PDF & Excel Report Generator
"""

import os
import tempfile
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font

from openpyxl.drawing.image import Image as XLImage

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image
)

from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

from io import BytesIO


# -----------------------------------------------------
# Save Plotly Figure
# -----------------------------------------------------

def save_chart(fig):

    temp = tempfile.NamedTemporaryFile(
        delete=False,
        suffix=".png"
    )

    fig.write_image(temp.name)

    return temp.name


# -----------------------------------------------------
# Excel Report
# -----------------------------------------------------

def create_excel_report(
        df,
        fig,
        sql_query,
        question):

    wb = Workbook()

    ws = wb.active

    ws.title = "Query Results"

    ws["A1"] = "AI SQL Analytics Report"

    ws["A1"].font = Font(
        size=18,
        bold=True
    )

    ws["A3"] = "Question"

    ws["B3"] = question

    ws["A4"] = "Generated"

    ws["B4"] = datetime.now().strftime(
        "%d-%m-%Y %H:%M"
    )

    ws["A6"] = "Generated SQL"

    ws["A7"] = sql_query

    start = 10

    for col, column in enumerate(df.columns, 1):

        ws.cell(
            row=start,
            column=col
        ).value = column

        ws.cell(
            row=start,
            column=col
        ).font = Font(
            bold=True
        )

    for r, row in enumerate(
            df.values.tolist(),
            start + 1):

        for c, value in enumerate(row, 1):

            ws.cell(
                row=r,
                column=c
            ).value = value

    if fig is not None:

        image_path = save_chart(fig)

        sheet = wb.create_sheet(
            "Visualization"
        )

        img = XLImage(image_path)

        img.width = 700

        img.height = 450

        sheet.add_image(
            img,
            "A1"
        )

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return output.getvalue()


# -----------------------------------------------------
# PDF Report
# -----------------------------------------------------

def create_pdf_report(
        df,
        fig,
        sql_query,
        question):

    output = BytesIO()

    pdf = SimpleDocTemplate(output)

    styles = getSampleStyleSheet()

    story = []

    story.append(
        Paragraph(
            "<b>AI SQL Analytics Report</b>",
            styles["Title"]
        )
    )

    story.append(
        Spacer(1, 12)
    )

    story.append(
        Paragraph(
            f"<b>Generated :</b> {datetime.now()}",
            styles["BodyText"]
        )
    )

    story.append(
        Paragraph(
            f"<b>Question :</b> {question}",
            styles["BodyText"]
        )
    )

    story.append(
        Spacer(1, 12)
    )

    story.append(
        Paragraph(
            "<b>Generated SQL</b>",
            styles["Heading2"]
        )
    )

    story.append(
        Paragraph(
            sql_query,
            styles["Code"]
        )
    )

    story.append(
        Spacer(1, 20)
    )

    if fig is not None:

        image_path = save_chart(fig)

        story.append(
            Paragraph(
                "<b>Visualization</b>",
                styles["Heading2"]
            )
        )

        story.append(
            Image(
                image_path,
                width=450,
                height=300
            )
        )

        story.append(
            Spacer(1, 20)
        )

    data = [df.columns.tolist()]

    data.extend(df.values.tolist())

    table = Table(data)

    table.setStyle(

        TableStyle([

            ("BACKGROUND", (0,0), (-1,0), colors.darkblue),

            ("TEXTCOLOR", (0,0), (-1,0), colors.white),

            ("GRID", (0,0), (-1,-1), 1, colors.black),

            ("BACKGROUND", (0,1), (-1,-1), colors.beige),

            ("BOTTOMPADDING", (0,0), (-1,0), 10)

        ])

    )

    story.append(table)

    pdf.build(story)

    output.seek(0)

    return output.getvalue()